# -*- coding: utf-8 -*-
"""生成简版硬件清单 Excel（BOM + 关键方案一句话）。自动算行高，保证每格完整显示。"""
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "硬件清单"

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr = PatternFill("solid", fgColor="1F4E79")
hdrf = Font(bold=True, color="FFFFFF")
GREEN = PatternFill("solid", fgColor="E2F0D9")
YEL = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")

# 列宽（单位≈字符数；加宽以少换行）
COLW = {"A": 4, "B": 16, "C": 40, "D": 7, "E": 38, "F": 26}
for col, w in COLW.items():
    ws.column_dimensions[col].width = w


def dispw(s):
    """显示宽度：中文/全角算 2，其余算 1。"""
    return sum(2 if ord(c) > 0x2E80 else 1 for c in str(s))


def nlines(s, width_units):
    """在给定列宽下这段文字大约占几行（按换行符 + 宽度估算）。"""
    if s is None or s == "":
        return 1
    total = 0
    for seg in str(s).split("\n"):
        total += max(1, math.ceil((dispw(seg) + 1) / max(1, width_units - 1)))
    return total


def set_h(row, cells, floor=22, factor=1.0):
    """cells = [(文本, 该格可用宽度单位)]；按最多行的格设行高，保证完整显示。

    factor<1：更保守地估宽（合并大格 Excel 不自动撑高，宁可算高一点、留白也别截字）。
    floor：行高下限。
    """
    ln = max((nlines(t, w * factor) for t, w in cells), default=1)
    ws.row_dimensions[row].height = max(floor, ln * 18 + 12)


def color(st):
    if st.startswith("✅"):
        return GREEN
    if st.startswith("🔄"):
        return YEL
    return GREY


r = 1
ws.cell(r, 1, "云小圈质检系统 · 硬件清单").font = Font(bold=True, size=14)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.row_dimensions[r].height = 24
r += 2

# ---------- BOM ----------
head = ["#", "硬件", "型号 / 规格", "数量", "作用", "状态"]
for c, h in enumerate(head, 1):
    cell = ws.cell(r, c, h)
    cell.fill = hdr; cell.font = hdrf; cell.border = border
    cell.alignment = Alignment("center", "center")
ws.row_dimensions[r].height = 22
r += 1

bom = [
    (1, "计算主机", "GPU RTX3080 / CUDA12.6 / Win", "1", "跑识别+大模型+服务端+MySQL", "✅ 已有"),
    (2, "上相机", "海康 MV-CS050-10GC（GigE 彩色 5MP）", "1", "拍上面(front)：颗粒+主控+SN", "✅ 已买"),
    (3, "下相机", "海康 MV-CS050-10GC", "1", "拍下面(back)：颗粒+PCB日期", "✅ 已买（软件待接第2台）"),
    (4, "镜头 ×2", "海康FA HV1050M-6MP 10-50mm手动变焦", "2", "成像；调好对焦后锁死", "✅ 已买"),
    (5, "光源（打光）", "条形/环形，可调亮度", "1~2", "让丝印清晰、压反光（第一杠杆）", "✅ 已买"),
    (6, "遮光布", "遮光布", "1", "遮环境杂光，稳定光照", "✅ 已买"),
    (7, "固定治具/机架", "上下相机对置、托盘居中", "1", "固定相机+光源+托盘，取景一致", "🔄 已固定未锁死（需防撞）"),
    (8, "托盘", "4 槽固定卡位", "若干", "放 4 根内存条", "✅ 已做（定制托盘在途）"),
    (9, "网络交换机", "GigE 千兆交换机", "1", "两相机→交换机→直连电脑", "✅ 已买"),
    (10, "触发方式", "静止即拍（纯软件）", "—", "放盘静止就自动拍", "🔄 待写软件"),
    (11, "SN 识别方式", "标签二维码解码（照片直接解 DataMatrix）", "—", "读 SN，精确、无需扫码枪硬件", "✅ 已实现"),
    (12, "指示灯", "可编程RGB灯（软件控色）×4", "4", "每槽一灯：合格绿/不合格红", "🔄 已买待到货"),
]
cols = "ABCDEF"
for row in bom:
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.border = border
        cell.alignment = Alignment("center" if c in (1, 4) else "left", "center", wrap_text=True)
    ws.cell(r, 6).fill = color(row[5])
    set_h(r, [(row[i], COLW[cols[i]]) for i in range(6)])
    r += 1

r += 1
# ---------- 关键方案（一句话）----------
ws.cell(r, 1, "关键方案（一句话）").font = Font(bold=True, size=12)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.row_dimensions[r].height = 22
r += 1
# 项目 = A:B 合并（够宽放"防重复放盘"），怎么做 = C:F 合并
ws.cell(r, 1, "项目").fill = hdr; ws.cell(r, 1).font = hdrf
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(r, 1).border = border; ws.cell(r, 1).alignment = Alignment("center", "center")
ws.cell(r, 3, "怎么做").fill = hdr; ws.cell(r, 3).font = hdrf
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
ws.cell(r, 3).border = border; ws.cell(r, 3).alignment = Alignment("center", "center")
for c in range(1, 7):
    ws.cell(r, c).border = border
ws.row_dimensions[r].height = 22
r += 1

NAMEW = COLW["A"] + COLW["B"]          # 项目列(A:B)可用宽
MERGEW = sum(COLW[c] for c in "CDEF")  # 怎么做列(C:F)可用宽
plan = [
    ("静止即拍", "放盘后画面不动 → 自动拍上下两面。纯软件，不用按钮/传感器。"),
    ("防重复放盘", "同一批里，新盘 4 根 SN 和之前测过的某盘一样 → 跳过不重复入库（读不到 SN 就比照片）。"),
    ("取盘灭灯", "拍完亮灯显示结果；检测到取盘（画面变空）→ 灭灯，放下一盘循环。"),
    ("指示灯", "可编程 RGB 灯，软件设色：合格🟢 / 不合格🔴 / 识别中🟡 / 空位⚫。"),
]
for k, v in plan:
    ws.cell(r, 1, k).font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1).border = border
    ws.cell(r, 1).alignment = Alignment("center", "center", wrap_text=True)
    cell = ws.cell(r, 3, v)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    cell.border = border
    cell.alignment = Alignment("left", "center", wrap_text=True)
    for c in range(1, 7):
        ws.cell(r, c).border = border
    # 合并大格保守估宽(factor)+保底2行高，确保长句完整显示不截字
    set_h(r, [(k, NAMEW), (v, MERGEW)], floor=46, factor=0.55)
    r += 1

ws.sheet_view.showGridLines = False

for name in ["硬件清单.xlsx", "硬件清单_v2.xlsx"]:
    try:
        wb.save(name)
        print("已生成", name)
        break
    except PermissionError:
        print(name, "被占用，换名重试...")
