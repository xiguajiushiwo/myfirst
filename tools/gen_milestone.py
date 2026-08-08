# -*- coding: utf-8 -*-
"""生成项目里程碑 Excel（当前工作模式 + 详细里程碑两页）。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ============ 单表：项目里程碑 ============
ws = wb.active
ws.title = "项目里程碑"
DONE, DOING, TODO, DEPRE = "✅ 完成", "🔄 进行中", "⏳ 待办", "🚫 弃用"
rows = [
    # ============ 一、模型时间（识别 + 判定 + 外观 + 大模型 + 测试）============
    ("一、模型时间（日期识别 · 综合判定 · 外观 · 大模型）",),
    ("", "PaddleOCR 引擎(GPU/server模型/整图+分块)", DONE, "PP-OCRv5"),
    ("", "日期解码 颗粒YWW(3位) / PCB·主控YYWW(4位)", DONE, "含单元测试；YYYYWW 入库"),
    ("", "整图固定模板：一次拍摄识别 4 根内存条(一张图)", DONE, "唯一识别路径；一盘4根同图 颗粒+每根PCB 全框，一次识别"),
    ("", "单根模板识别(按型号固定框)", DEPRE, "已弃用——只放4根一盘，只用整图固定模板；旧模板JSON仅留档"),
    ("", "规则识别(整图自动找，无模板兜底)", DONE, "会漏检，仅临时用"),
    ("", "多方案提准(自动翻正180°+CLAHE增强+字符挽救)", DONE, "治倒装/低对比/点阵字误读"),
    ("", "逐颗大模型兜底(低置信转VL，保留OCR真实读数/置信)", DONE, "透明不掩盖，读错也照显"),
    ("", "PCB日期集成到反面(4位框，低置信转大模型)", DONE, "2536→25年36周"),
    ("", "整图大模型核对漏检(开关：VL整图数颗粒比对)", DONE, "比OCR多则告警疑似漏检"),
    ("", "综合判定：单灯·逐颗比较·严禁多数表决·定位第N颗", DONE, "防偷换单颗；盲点单列提示人工；阈值(周差)前端可调"),
    ("", "外观质检 Qwen-VL 三项(元器件/金手指/二维码标记)", DONE, "三态bool + 不合格逐条说明"),
    ("", "多模态大模型：多provider配置+一键切换+用量统计", DONE, "OpenAI兼容，可接多个"),
    ("", "大模型自动降级(第一个失败自动切下一个)", DONE, "按启用顺序逐个尝试，全失败才报错"),
    ("", "提示词独立文件(便于调词)", DONE, "app/prompts.py"),
    ("", "单元测试(pytest 27例：解码+判定铁律+空位+拆根)", DONE, "锁住判定链与拆分正确性"),
    ("", "准确率评测脚本(给标注答案算准确率)", DONE, "tools/eval_accuracy"),
    ("", "识别准确率", DOING, "清晰图~99%/差图~88%；瓶颈=拍摄清晰度(镜头/打光/对焦)"),
    ("", "模板批量自动生成工具", DOING, "已有OCR自动框；独立工具待完善"),
    ("", "本地多模态模型(离线/省按次费用)", TODO, "长期自用/商业化建议"),

    # ============ 二、硬件（主机 + 触发 + 部署运维）============
    ("二、硬件（计算主机 · 触发 · 部署运维）",),
    ("", "计算主机 GPU(RTX3080 / CUDA12.6)", DONE, "本地跑 PaddleOCR"),
    ("", "分级滚动日志 + 健康检查 + 运行指标接口", DONE, "/api/health /api/metrics"),
    ("", "进程守护(服务/开机自启脚本)", DONE, "崩溃自拉起"),
    ("", "拍照触发装置(按钮/托盘到位开关/定时)", TODO, "三选一，待现场定；最省事=工位屏大按钮"),
    ("", "部署 runbook / Docker 一键部署", TODO, "GPU+海康SDK+paddle 换机重装"),
    ("", "登录 + 角色鉴权", TODO, "内网暂缓，交付前收口"),

    # ============ 三、相机 ============
    ("三、相机（海康 MV-CS050-10GC）",),
    ("", "MVS SDK 驱动(枚举/取流/抓帧存图/转BGR)", DONE, "实测通"),
    ("", "实时预览 MJPEG 流", DONE, "工作台实时画面"),
    ("", "曝光/增益 代码+接口+前端可调", DONE, "清晰度不够先调曝光+打光"),
    ("", "拍照 → 识别 一条龙", DONE, "拍正反 → 一键识别"),
    ("", "双相机 上/下 各一台(海康MV-CS050-10GC GigE彩色5MP)", DOING, "硬件已定；代码仅接1台，待接第2台并按相机SN绑定上/下"),
    ("", "相机写文件夹 → 自动监听识别", DOING, "FolderWatchFeeder 已具备；接真机模板后即用于托盘"),

    # ============ 四、镜头（光学：镜头/打光/对焦）============
    ("四、镜头（光学：选型 · 打光 · 对焦）",),
    ("", "镜头选型：海康FA HV1050M-6MP(10-50mm手动变焦,600万)", DONE, "已定 ×2；现场调好视场/对焦后锁死"),
    ("", "打光(斜射/环形，让激光丝印立体清晰)", TODO, "准确率第一杠杆，硬件轨"),
    ("", "对焦 / 景深(4根同一平面全清晰)", TODO, "HV1050M 手动变焦现场调好锁死"),

    # ============ 五、模具（治具 / 固定架 / 定位）============
    ("五、模具（治具 · 固定架 · 定位）",),
    ("", "固定治具：相机+托盘+光源位置固定，取景一致", TODO, "取景固定=模板一劳永逸的前提"),
    ("", "托盘定位(卡位重复精度，每次放回同一位置)", TODO, "位置漂移会让固定框错位"),

    # ============ 六、托盘（上料方式 + 托盘相关识别）============
    ("六、托盘（上料方式 · 空位 · 拆根）",),
    ("", "托盘上料：一盘4根固定位，每盘拍照", DONE, "改动：弃用电机传送/弹仓/到位握手，改托盘"),
    ("", "正反面 = 双相机各拍一面(不翻盘)", DONE, "定稿；正背按托盘位配对"),
    ("", "空位检测：先数数量(哪些槽有条)再识别", DONE, "边缘密度判空/满，空槽跳过并标「空位」，不误判"),
    ("", "一盘拆成 N 条独立记录", DONE, "每根逐颗判定(不跨根)/各自SN/各自外观/各自归批，slot_pos入库"),
    ("", "工位屏 /station：一盘逐根出合格/不合格 + 本盘良率", DONE, "全屏零点击；托盘拍完自动刷"),
    ("", "按真实托盘取景建4槽整图模板(正/反两套)", TODO, "待你双相机+托盘真实照片"),
    ("", "占位阈值标定(满盘/空盘各拍一张)", TODO, "tools/calibrate_slots.py 已就绪，待真机标定"),
    ("", "防重复入库(同盘拍两次幂等)", TODO, "按SN+时间/图像去重"),

    # ============ 七、数据（记录 + 追溯 + 看板 + 报表 + 对接）============
    ("七、数据（记录 · 追溯 · 看板 · 报表 · 对接）",),
    ("", "MySQL 质检记录(含批次/图片/slot_pos/审计字段)", DONE, "每根一条，SN 不唯一留历史"),
    ("", "增量迁移(改表不丢历史数据)", DONE, ""),
    ("", "原图+标注图归档并绑定记录(可回看举证)", DONE, "archive/ 永久区"),
    ("", "审计日志(复查/删除留痕) + 数据库定时备份", DONE, ""),
    ("", "批次登记(客户/品牌/容量/频率/品相/来源/单号)", DONE, "每根继承、只读SN"),
    ("", "良品率看板(今日/累计/按客户/按批次) + 人工复查", DONE, "/manage"),
    ("", "工作台 / 管理看板 / 系统设置 界面", DONE, "统一设计，顶栏+卡片"),
    ("", "SN 读取上扫码枪(替代大模型读标签)", TODO, "SN 是追溯主键，需可信；大模型仅兜底"),
    ("", "按 SN 追溯举证(输SN调出历次记录+当时图)", TODO, "数据已齐，缺查询页"),
    ("", "同 SN 历史比对(防偷换第二道：颗粒日期变了告警)", TODO, "同根多次质检对比"),
    ("", "来源/品相 质量画像(按渠道/品相不良率)", TODO, "指导进货"),
    ("", "报表导出 Excel/PDF(对账 / 质检报告)", TODO, ""),
    ("", "不合格品处置闭环(返工/隔离/报废 状态流转)", TODO, ""),
    ("", "与交易系统(云小圈)打通：质检结果回流", TODO, ""),
]
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
grp_fill = PatternFill("solid", fgColor="D6E4F5")
grp_font = Font(bold=True, size=11, color="1F4E79")
fills = {DONE: PatternFill("solid", fgColor="E2F0D9"),
         DOING: PatternFill("solid", fgColor="FFF2CC"),
         TODO: PatternFill("solid", fgColor="F2F2F2"),
         DEPRE: PatternFill("solid", fgColor="E7E6E6")}
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(["模块", "里程碑 / 子项", "状态", "详细说明"])
for cc in ws[1]:
    cc.fill = hdr_fill; cc.font = hdr_font
    cc.alignment = Alignment("center", "center"); cc.border = border
r = 2
for row in rows:
    if len(row) == 1:
        ws.append([row[0], "", "", ""])
        ws.merge_cells(f"A{r}:D{r}")
        cell = ws.cell(r, 1)
        cell.fill = grp_fill; cell.font = grp_font
        cell.alignment = Alignment("left", "center")
        for col in range(1, 5):
            ws.cell(r, col).border = border
    else:
        _, item, st, note = row
        ws.append(["", item, st, note])
        ws.cell(r, 3).fill = fills[st]
        ws.cell(r, 3).alignment = Alignment("center", "center")
        for col in range(1, 5):
            ws.cell(r, col).border = border
            ws.cell(r, col).alignment = Alignment(vertical="center", wrap_text=(col in (2, 4)))
    r += 1
done = sum(1 for x in rows if len(x) == 4 and x[2] == DONE)
doing = sum(1 for x in rows if len(x) == 4 and x[2] == DOING)
todo = sum(1 for x in rows if len(x) == 4 and x[2] == TODO)
tot = done + doing + todo
ws.append([])
ws.append(["汇总", f"完成 {done} · 进行中 {doing} · 待办 {todo}（共 {tot} 项）",
           f"{done / tot * 100:.0f}%", "截至 2026-07-15"])
for col in range(1, 5):
    ws.cell(r + 1, col).font = Font(bold=True)
    ws.cell(r + 1, col).fill = PatternFill("solid", fgColor="FCE4D6")
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 11
ws.column_dimensions["D"].width = 60
ws.freeze_panes = "A2"
ws.sheet_view.showGridLines = False

for name in ["云小圈质检系统_项目里程碑_v3.xlsx", "云小圈质检系统_项目里程碑_v3b.xlsx"]:
    try:
        wb.save(name)
        print("已生成", name, f"| 完成{done} 进行中{doing} 待办{todo} 共{tot} ({done/tot*100:.0f}%)")
        break
    except PermissionError:
        print(name, "被占用(Excel开着)，换名重试...")
