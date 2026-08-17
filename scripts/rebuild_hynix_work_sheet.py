from pathlib import Path
from shutil import copy2

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TARGET = ROOT / "海力士内存条质检工作表_进度版_成像版_2026-08-12.xlsx"
BACKUP = ROOT / "海力士内存条质检工作表_进度版_2026-08-12_原版备份.xlsx"

# Keep the three rows already revised by the user exactly as written.
ROWS = [
    ["一、外观检测", "金手指检查", "识别划痕、氧化、污渍和烧坏，并标出位置。", 10, "当前日期识别流程未启用外观质检；需补充金手指缺陷样本，先做划痕、氧化和烧坏分类，再接入复核。"],
    ["一、外观检测", "PCB 与芯片检查", "识别 PCB 划痕、烧焦、破损及芯片崩角、烧坏。", 10, "当前主要完成日期识别；需增加 PCB、主控和存储芯片区域，并采集正常与缺陷样本验证。"],
    ["一、外观检测", "元器件掉件", "输出缺失、偏移或损坏元器件的具体点位。", 5, "尚无点位级掉件检测；需先标定海力士版型点位，再用几何对齐加 ROI 分类或异常检测实现。"],
    ["二、成像", "反光与颜色", "拍摄图像无明显反光，元器件、芯片、金手指和日期等清晰可见。", 80, "最新实拍图整体光照均匀，芯片丝印、日期、金手指和小元器件清晰；仅托盘边缘及少量金属表面有局部高光。"],
    ["三、识别能力", "日期识别", "分别输出存储芯片、主控和 PCB 日期。", 80, "当前可分区识别存储芯片、主控和 PCB 日期，并进行日期一致性判断；海力士日期格式还需继续校准。"],
    ["三、识别能力", "识别速度", "一盘内存条识别不超过 13 秒，并记录拍照、裁图、OCR、标签识别、归档和入库等分段耗时。", 50, "分段耗时记录已经实现；现有实测最快约 12 秒，2000w像素相机尚待测试。"],
    ["三、识别能力", "模型消耗", "记录每次调用次数及输入、输出和总 Token。", 60, "耗时和 Token 记录字段已经接入；当前外观模型停用，实际消耗仍需在启用模型后连续测试。"],
    ["四、电脑与服务器配置", "RTX 3080 服务器", "RTX 3080 可稳定承担日期识别、服务端和并发任务，显存使用不导致任务中断。", 90, "已配置 RTX 3080；当前用于 PaddleOCR、服务端和数据库，需用海力士整盘连续运行确认显存峰值。"],
    ["四、电脑与服务器配置", "CUDA 与 GPU 推理环境", "CUDA、显卡驱动与 GPU 版 PaddlePaddle 匹配，识别任务可使用 GPU。", 85, "项目环境已按 CUDA 12.6 和 paddlepaddle-gpu 配置；还需在新相机和海力士样本上确认实际 GPU 推理。"],
    ["四、电脑与服务器配置", "Python 与 FastAPI 服务", "服务可正常启动，识别、订单、记录和看板接口可用。", 90, "Python 3.12、FastAPI 和主要接口已运行；仍需将海力士整条链路做一次连续业务验证。"],
    ["四、电脑与服务器配置", "MySQL 数据库", "质检记录、订单、SN、耗时和图片路径可持久化，并可按订单查询。", 85, "MySQL 质检记录和订单查询已接入；订单绑定数据仍需通过真实订单质检补齐验证。"],
    ["四、电脑与服务器配置", "大模型接口配置", "接口密钥、模型名称和超时配置独立管理，调用失败不影响本地日期识别。", 70, "模型接口配置已存在，日期识别以本地 OCR 为主；外观模型当前停用，需补充失败重试和降级验证。"],
    ["四、电脑与服务器配置", "硬盘与图片存储", "原图、裁剪图和识别结果按订单、槽位、SN 归档，路径可追溯。", 75, "图片归档代码已实现，项目已有 uploads、archive 和 outputs 目录；需按真实订单检查长期存储和空间告警。"],
    ["四、电脑与服务器配置", "网络与设备通信", "服务器与两台 GigE 相机、数据库及金蝶接口通信稳定。", 80, "相机和服务接口已具备通信能力；需持续运行验证断线重连、1Gbps 网络和金蝶接口异常处理。"],
    ["二、成像", "正反面相机", "两台海康相机均在线，正反面图像清晰且能完成同步取图。", 85, "健康接口曾确认两台相机在线；正面已按序列号重新绑定，仍需用海力士样品确认清晰度。"],
    ["五、相机与硬件工装", "相机序列号绑定", "正面和反面角色固定绑定到指定序列号，重启后不串 camera。", 95, "正面 DB1224590、反面 DB1623157 已配置绑定并可识别；需现场重启后再做一次确认。"],
    ["二、成像", "镜头与成像范围", "镜头焦距、焦点和视野覆盖整盘，芯片日期与金手指细节可辨。", 85, "已配置海康 FA HV1050M-6MP、10–50mm 镜头；当前成像基本满足识别，海力士版型仍需微调焦距。"],
    ["二、成像", "光源与遮光", "光照均匀，芯片丝印、日期和金手指无影响识别的高光或阴影。", 75, "已有补光和遮光条件，最新实拍整体清晰；少量金属表面仍有高光，需继续调光并做多批次验证。"],
    ["五、相机与硬件工装", "相机架与固定治具", "相机、镜头、光源和托盘位置固定，重复拍摄不偏移。", 20, "架体已固定但锁紧和防碰撞仍需现场确认；需标记位置、加固并做重复拍摄对齐测试。"],
    ["五、相机与硬件工装", "托盘与槽位", "4 个槽位位置固定，内存条方向一致，每槽可独立裁图和出结果。", 65, "已有 4 槽托盘、槽位检测和裁图代码；需按海力士版型重新标定并用满盘实拍验证。"],
    ["五、相机与硬件工装", "静止即拍触发", "托盘放稳后自动触发正反面拍照，避免移动中取图和重复识别。", 75, "静止检测、双相机取图和目录监听流程已有；需现场连续放盘验证误触发、漏触发和断线重试。"],
    ["六、软件流程与数据", "选择订单开始质检", "开始质检前选择一个采购订单，后续照片和记录自动归属于该订单。", 60, "订单 ID 绑定代码已支持；需用真实采购订单跑通从选择、拍照到入库的完整链路。"],
    ["六、软件流程与数据", "一盘四槽识别入库", "一盘识别后按槽位拆分为独立记录，每根关联 SN、日期、结果和图片。", 80, "整盘取图、槽位裁剪和独立记录流程已有；需用海力士 4 槽样品核对槽位顺序和数据完整性。"],
    ["六、软件流程与数据", "订单、SN 与图片追溯", "可按订单查看整批记录，也可按 SN 查看历史图片和识别结果。", 70, "订单查询、SN 历史和图片路径已实现；现有历史数据订单绑定不足，需从新订单开始完整积累。"],
    ["六、软件流程与数据", "订单查询与供应商筛选", "支持完整或部分订单号搜索，并可按供应商筛选有效采购订单。", 100, "订单号搜索和供应商筛选已完成并验证。"],
    ["六、软件流程与数据", "质检看板与质检员权限", "普通质检员可查看订单进度、问题、待复查记录和批次结果。", 95, "质检看板和普通质检员访问权限已完成；需用普通质检员账号做一次完整业务验证。"],
    ["六、软件流程与数据", "金蝶增量同步", "前端显示上次同步时间、成功或失败状态，并可执行增量同步。", 80, "增量同步已有成功记录，前端状态字段已接入；需继续观察定时任务是否按间隔稳定执行。"],
    ["六、软件流程与数据", "金蝶全量同步", "可手动执行全量同步，显示开始时间、结束时间、导入数量和失败原因。", 40, "全量接口和状态字段已有，但尚无成功记录；需执行一次全量同步并处理超时、重复订单和失败提示。"],
    ["六、软件流程与数据", "日志与异常处理", "相机、OCR、同步、数据库和归档失败时有日志、错误提示和可定位阶段。", 75, "分阶段耗时、同步错误和连续失败字段已有；需补齐前端异常提示和现场故障复现测试。"],
    ["六、软件流程与数据", "报表与结果导出", "可按订单导出识别结果、问题位置、复查状态和图片索引。", 30, "订单和记录数据已具备，专用导出功能尚未完整验证；需增加按订单导出 Excel/CSV 和问题清单。"],
]


def build():
    if TARGET.exists() and not BACKUP.exists():
        copy2(TARGET, BACKUP)
    wb = Workbook()
    ws = wb.active
    ws.title = "完成进度"
    ws.merge_cells("A1:E1")
    ws["A1"] = "海力士内存条质检项目完成进度"
    ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:E2")
    ws["A2"] = "日期：2026-08-12；完成度为当前项目实际状态估算，未上线的检测能力不按预留代码计入。"
    ws["A2"].font = Font(name="Microsoft YaHei", size=11, color="404040")
    ws["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A2"].alignment = Alignment(vertical="center")
    headers = ["分类", "检查事项", "完成标准", "当前完成度", "当前做到的程度 / 未完成实施方案"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(4, col, value)
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    colors = {"一、外观检测": "FCE4D6", "二、成像": "D9EAD3", "三、识别能力": "FFF2CC", "四、电脑与服务器配置": "DDEBF7", "五、相机与硬件工装": "E2F0D9", "六、软件流程与数据": "E4DFEC"}
    thin = Side(style="thin", color="B7C9DA")
    order = {category: index for index, category in enumerate(colors)}
    sorted_rows = sorted(ROWS, key=lambda row: order[row[0]])
    for row_num, values in enumerate(sorted_rows, 5):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_num, col, value)
            cell.font = Font(name="Microsoft YaHei", size=10, bold=col == 1)
            cell.alignment = Alignment(horizontal="center" if col in (1, 4) else "left", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if col == 1:
                cell.fill = PatternFill("solid", fgColor=colors[values[0]])
            if col == 4:
                cell.number_format = '0"%"'
        ws.row_dimensions[row_num].height = 58
    for col, width in enumerate([22, 23, 45, 14, 66], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    last_row = 4 + len(ROWS)
    ws.conditional_formatting.add(f"D5:D{last_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="5B9BD5"))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{last_row}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:E{last_row}"
    wb.save(TARGET)
    print(f"saved: {TARGET}")
    print(f"rows: {len(ROWS)}")


if __name__ == "__main__":
    build()
